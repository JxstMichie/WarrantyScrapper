from openpyxl import load_workbook
import scrape
import tkinter
"""
Test Serials:
DELL -> JG9LBB4
Lenovo -> PF2ZXCS1
HP -> 5CD328GMQB
"""

serial = input()


def write_device_to_excel(device:dict):

    wb = load_workbook("Devices.xlsm", keep_vba=True)
    ws = wb.active

    next_row = 2 
    while ws[f"A{next_row}"].value not in (None, ""):
        next_row += 1

    # Write the data into Excel
    ws[f"A{next_row}"] = device["serial_number"]
    ws[f"B{next_row}"] = device["brand"]
    ws[f"C{next_row}"] = device["device_name"]
    ws[f"D{next_row}"] = device["warranty_end_date"]


    wb.save("Devices.xlsm")
    print(
        f"Added {device['serial_number']} | {device['brand']} | {device['device_name']} | {device['warranty_end_date']} to row {next_row}"
    )


def main():
    device = scrape.scrape_serial(serial)
    write_device_to_excel(device)

main()
