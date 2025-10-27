import customtkinter as ctk
from openpyxl import load_workbook
import scrape
import getpass

class DeviceTicketSystem(ctk.CTk):
    def __init__(self):
        super().__init__()
        
        self.title("Device Ticket System")
        self.geometry("1200x700")
        
        # Set light theme by default
        ctk.set_appearance_mode("light")
        ctk.set_default_color_theme("blue")
        
        # Get Windows username
        self.current_user = getpass.getuser()
        
        # Sidebar state
        self.sidebar_expanded = False
        self.sidebar_width_collapsed = 60
        self.sidebar_width_expanded = 200
        
        # Cache for loaded tickets
        self.tickets_cache = []
        self.cache_valid = False
        
        # Main layout
        self.grid_columnconfigure(1, weight=1)
        self.grid_rowconfigure(0, weight=1)
        
        # Sidebar
        self.create_sidebar()
        
        # Main content area
        self.main_frame = ctk.CTkFrame(self, fg_color="#F8F9FA", corner_radius=0)
        self.main_frame.grid(row=0, column=1, sticky="nsew")
        
        # Show default mode
        self.show_view_mode()
    
    def create_sidebar(self):
        self.sidebar = ctk.CTkFrame(self, width=self.sidebar_width_collapsed, corner_radius=0, fg_color="white")
        self.sidebar.grid(row=0, column=0, sticky="nsew")
        self.sidebar.grid_rowconfigure(5, weight=1)
        self.sidebar.grid_propagate(False)
        
        # Bind hover events
        self.sidebar.bind("<Enter>", self.expand_sidebar)
        self.sidebar.bind("<Leave>", self.collapse_sidebar)
        
        title = ctk.CTkLabel(
            self.sidebar, 
            text="📱",
            font=ctk.CTkFont(size=24),
            text_color="#1F2937"
        )
        title.grid(row=0, column=0, padx=0, pady=(30, 40))
        title.bind("<Enter>", self.expand_sidebar)
        
        self.title_text = ctk.CTkLabel(
            self.sidebar, 
            text="Task app",
            font=ctk.CTkFont(size=18, weight="bold"),
            text_color="#1F2937"
        )
        
        # Mode buttons with icons
        self.view_btn = ctk.CTkButton(
            self.sidebar, 
            text="📋",
            command=self.show_view_mode,
            height=35,
            width=40,
            font=ctk.CTkFont(size=18),
            fg_color="#E5E7EB",
            text_color="#374151",
            hover_color="#D1D5DB",
            corner_radius=6
        )
        self.view_btn.grid(row=1, column=0, padx=10, pady=5)
        self.view_btn.bind("<Enter>", self.expand_sidebar)
        
        self.lookup_btn = ctk.CTkButton(
            self.sidebar,
            text="🔍",
            command=self.show_lookup_mode,
            height=35,
            width=40,
            font=ctk.CTkFont(size=18),
            fg_color="transparent",
            text_color="#6B7280",
            hover_color="#F3F4F6",
            corner_radius=6
        )
        self.lookup_btn.grid(row=2, column=0, padx=10, pady=5)
        self.lookup_btn.bind("<Enter>", self.expand_sidebar)
        
        self.add_btn = ctk.CTkButton(
            self.sidebar,
            text="➕",
            command=self.show_add_mode,
            height=35,
            width=40,
            font=ctk.CTkFont(size=18),
            fg_color="transparent",
            text_color="#6B7280",
            hover_color="#F3F4F6",
            corner_radius=6
        )
        self.add_btn.grid(row=3, column=0, padx=10, pady=5)
        self.add_btn.bind("<Enter>", self.expand_sidebar)
        
        # User info
        self.user_icon = ctk.CTkLabel(
            self.sidebar,
            text="👤",
            font=ctk.CTkFont(size=18),
            text_color="#6B7280"
        )
        self.user_icon.grid(row=6, column=0, padx=0, pady=20)
        self.user_icon.bind("<Enter>", self.expand_sidebar)
        
        self.user_text = ctk.CTkLabel(
            self.sidebar,
            text=f"User: {self.current_user}",
            font=ctk.CTkFont(size=12),
            text_color="#6B7280"
        )
    
    def expand_sidebar(self, event=None):
        if self.sidebar_expanded:
            return
        self.sidebar_expanded = True
        self.sidebar.configure(width=self.sidebar_width_expanded)
        self._show_expanded_content()
    
    def collapse_sidebar(self, event=None):
        if not self.sidebar_expanded:
            return
        self.sidebar_expanded = False
        self.sidebar.configure(width=self.sidebar_width_collapsed)
        self._hide_expanded_content()
    
    def _show_expanded_content(self):
        self.title_text.grid(row=0, column=1, padx=(0, 20), pady=(30, 40), sticky="w")
        self.view_btn.configure(text="📋 View Devices", width=180)
        self.view_btn.grid(padx=15, sticky="ew")
        self.lookup_btn.configure(text="🔍 Look Up", width=180)
        self.lookup_btn.grid(padx=15, sticky="ew")
        self.add_btn.configure(text="➕ Add Device", width=180)
        self.add_btn.grid(padx=15, sticky="ew")
        self.user_text.grid(row=6, column=1, padx=(0, 20), pady=20, sticky="w")
    
    def _hide_expanded_content(self):
        self.title_text.grid_forget()
        self.view_btn.configure(text="📋", width=40)
        self.view_btn.grid(padx=10, sticky="")
        self.lookup_btn.configure(text="🔍", width=40)
        self.lookup_btn.grid(padx=10, sticky="")
        self.add_btn.configure(text="➕", width=40)
        self.add_btn.grid(padx=10, sticky="")
        self.user_text.grid_forget()
    
    def update_active_button(self, active_btn):
        for btn in [self.view_btn, self.lookup_btn, self.add_btn]:
            btn.configure(fg_color="transparent", text_color="#6B7280")
        active_btn.configure(fg_color="#E5E7EB", text_color="#374151")
    
    def clear_main_frame(self):
        for widget in self.main_frame.winfo_children():
            widget.destroy()
    
    # --------------- View Mode -----------------
    def show_view_mode(self):
        self.clear_main_frame()
        self.update_active_button(self.view_btn)
        
        # Header with search
        header = ctk.CTkFrame(self.main_frame, fg_color="transparent", height=80)
        header.pack(fill="x", padx=30, pady=(20, 10))
        header.pack_propagate(False)
        
        # Search bar at top
        search_container = ctk.CTkFrame(header, fg_color="transparent")
        search_container.pack(fill="x", pady=(0, 15))
        
        self.search_var = ctk.StringVar()
        self.search_var.trace("w", self.filter_tickets)
        
        search_entry = ctk.CTkEntry(
            search_container,
            placeholder_text="🔍 Search...",
            width=300,
            height=38,
            textvariable=self.search_var,
            fg_color="white",
            border_color="#E5E7EB",
            text_color="#1F2937"
        )
        search_entry.pack(side="left")
        
        # Title below search
        title = ctk.CTkLabel(
            header,
            text="Device Tickets",
            font=ctk.CTkFont(size=22, weight="bold"),
            text_color="#1F2937",
            anchor="w"
        )
        title.pack(side="left", anchor="w")
        
        # Table container
        table_container = ctk.CTkFrame(self.main_frame, fg_color="white", corner_radius=8)
        table_container.pack(fill="both", expand=True, padx=30, pady=(10, 30))
        
        # Table header
        header_frame = ctk.CTkFrame(table_container, fg_color="transparent", height=45)
        header_frame.pack(fill="x", padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        headers = [("Serial", 0.2), ("Brand", 0.18), ("Device", 0.32), ("Warranty", 0.15), ("Author", 0.15)]
        left_padding = 20
        
        for i, (text, width) in enumerate(headers):
            x_pos = left_padding + sum(w * 1100 for _, w in headers[:i])
            lbl = ctk.CTkLabel(
                header_frame,
                text=text,
                font=ctk.CTkFont(size=12, weight="normal"),
                text_color="#6B7280",
                anchor="w",
                width=int(width * 1100)
            )
            lbl.place(x=x_pos, rely=0.5, anchor="w")
        
        # Separator line
        separator = ctk.CTkFrame(table_container, fg_color="#E5E7EB", height=1)
        separator.pack(fill="x", padx=0)
        
        # Scrollable content
        self.tickets_frame = ctk.CTkScrollableFrame(
            table_container,
            fg_color="white",
            scrollbar_button_color="#D1D5DB",
            scrollbar_button_hover_color="#9CA3AF"
        )
        self.tickets_frame.pack(fill="both", expand=True, padx=0, pady=0)
        
        self.load_tickets()
    
    def load_tickets(self):
        for widget in self.tickets_frame.winfo_children():
            widget.destroy()
        
        # Use cache if valid
        if self.cache_valid and self.tickets_cache:
            for ticket_data in self.tickets_cache:
                self.create_ticket_row(*ticket_data)
            return
        
        # Load from file and update cache
        try:
            wb = load_workbook("Devices.xlsm", keep_vba=True, data_only=True)
            ws = wb.active
            
            self.tickets_cache = []
            row = 2
            while ws[f"A{row}"].value not in (None, ""):
                serial = ws[f"A{row}"].value
                brand = ws[f"B{row}"].value
                device_name = ws[f"C{row}"].value
                warranty = ws[f"D{row}"].value
                author = ws[f"E{row}"].value if ws[f"E{row}"].value else "Unknown"
                
                ticket_data = (serial, brand, device_name, warranty, author)
                self.tickets_cache.append(ticket_data)
                self.create_ticket_row(*ticket_data)
                row += 1
            
            wb.close()
            self.cache_valid = True
        except Exception as e:
            ctk.CTkLabel(
                self.tickets_frame,
                text=f"Error loading tickets: {str(e)}",
                text_color="red"
            ).pack(pady=20)
    
    def create_ticket_row(self, serial, brand, device_name, warranty, author):
        row = ctk.CTkFrame(self.tickets_frame, fg_color="transparent", height=50)
        row.pack(fill="x", padx=0, pady=4)
        row.pack_propagate(False)
        
        fields = [(serial, 0.2), (brand, 0.18), (device_name, 0.32), (warranty, 0.15), (author, 0.15)]
        left_padding = 20
        
        for i, (value, width) in enumerate(fields):
            x_pos = left_padding + sum(w * 1100 for _, w in fields[:i])
            lbl = ctk.CTkLabel(
                row,
                text=str(value),
                font=ctk.CTkFont(size=13),
                text_color="#1F2937" if i < 2 else "#6B7280",
                anchor="w",
                width=int(width * 1100)
            )
            lbl.place(x=x_pos, rely=0.5, anchor="w")
        
        # Bottom border
        border = ctk.CTkFrame(row, fg_color="#F3F4F6", height=1)
        border.pack(side="bottom", fill="x", padx=20)
    
    def filter_tickets(self, *args):
        search_term = self.search_var.get().lower()
        for row in self.tickets_frame.winfo_children():
            row_text = ""
            for widget in row.winfo_children():
                if isinstance(widget, ctk.CTkLabel):
                    row_text += widget.cget("text").lower() + " "
            if search_term in row_text:
                row.pack(fill="x", padx=0, pady=4)
            else:
                row.pack_forget()
    
    # --------------- Lookup Mode -----------------
    def show_lookup_mode(self):
        self.clear_main_frame()
        self.update_active_button(self.lookup_btn)
        
        # Main container
        main_container = ctk.CTkFrame(self.main_frame, fg_color="transparent")
        main_container.pack(fill="both", expand=True, padx=30, pady=20)
        
        # Search input container
        input_container = ctk.CTkFrame(main_container, fg_color="white", corner_radius=12)
        input_container.pack(fill="x", pady=(0, 20))
        
        title = ctk.CTkLabel(
            input_container, 
            text="Look Up Device", 
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color="#1F2937"
        )
        title.pack(pady=(30, 20))
        
        # Brand selector
        brand_frame = ctk.CTkFrame(input_container, fg_color="transparent")
        brand_frame.pack(pady=10, padx=40, fill="x")
        
        brand_inner = ctk.CTkFrame(brand_frame, fg_color="transparent")
        brand_inner.pack(side="left", fill="x", expand=True)
        
        ctk.CTkLabel(
            brand_inner, 
            text="Brand:", 
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#374151"
        ).pack(anchor="w", pady=(0, 8))
        self.brand_selector = ctk.CTkOptionMenu(
            brand_inner, 
            values=["HP", "Dell", "Lenovo"], 
            height=40,
            width=200,
            fg_color="#3B82F6",
            button_color="#2563EB",
            button_hover_color="#1D4ED8",
            text_color="white",
            dropdown_fg_color="white",
            dropdown_hover_color="#EFF6FF",
            dropdown_text_color="#1F2937",
            font=ctk.CTkFont(size=14, weight="bold"),
            corner_radius=8
        )
        self.brand_selector.pack(anchor="w")
        
        # Serial number input
        serial_frame = ctk.CTkFrame(input_container, fg_color="transparent")
        serial_frame.pack(pady=10, padx=40, fill="x")
        
        serial_inner = ctk.CTkFrame(serial_frame, fg_color="transparent")
        serial_inner.pack(side="left", fill="x", expand=True)
        
        ctk.CTkLabel(
            serial_inner, 
            text="Serial Number:", 
            font=ctk.CTkFont(size=14, weight="bold"),
            text_color="#374151"
        ).pack(anchor="w", pady=(0, 8))
        self.lookup_serial = ctk.CTkEntry(
            serial_inner, 
            placeholder_text="Enter serial number", 
            height=40,
            width=400,
            fg_color="white",
            border_color="#D1D5DB",
            border_width=2,
            text_color="#1F2937",
            placeholder_text_color="#9CA3AF",
            corner_radius=8,
            font=ctk.CTkFont(size=14)
        )
        self.lookup_serial.pack(anchor="w")
        
        # Search button
        lookup_btn = ctk.CTkButton(
            input_container, 
            text="Search Device", 
            height=45, 
            width=200,
            command=self.perform_lookup,
            fg_color="#3B82F6",
            hover_color="#2563EB",
            corner_radius=8,
            font=ctk.CTkFont(size=15, weight="bold")
        )
        lookup_btn.pack(pady=(20, 30), padx=40, anchor="w")
        
        # Results container
        self.results_container = ctk.CTkFrame(main_container, fg_color="white", corner_radius=8)
        self.results_container.pack(fill="both", expand=True)
        
        # Initial empty state
        self.show_empty_results()
    
    def show_empty_results(self):
        for widget in self.results_container.winfo_children():
            widget.destroy()
        
        empty_label = ctk.CTkLabel(
            self.results_container,
            text="Enter a serial number and select a brand to search",
            font=ctk.CTkFont(size=14),
            text_color="#9CA3AF"
        )
        empty_label.place(relx=0.5, rely=0.5, anchor="center")
    
    def perform_lookup(self):
        brand = self.brand_selector.get()
        serial = self.lookup_serial.get().strip()
        
        if not serial:
            self.show_error_results("Please enter a serial number")
            return
        
        # Clear results container
        for widget in self.results_container.winfo_children():
            widget.destroy()
        
        # Search in Excel file
        try:
            wb = load_workbook("Devices.xlsm", keep_vba=True, data_only=True)
            ws = wb.active
            
            device_found = False
            row = 2
            while ws[f"A{row}"].value not in (None, ""):
                file_serial = str(ws[f"A{row}"].value).strip()
                file_brand = str(ws[f"B{row}"].value).strip()
                
                if file_serial.lower() == serial.lower() and file_brand.lower() == brand.lower():
                    # Device found
                    device_name = ws[f"C{row}"].value
                    warranty = ws[f"D{row}"].value
                    author = ws[f"E{row}"].value if ws[f"E{row}"].value else "Unknown"
                    
                    self.display_device_result(file_serial, file_brand, device_name, warranty, author)
                    device_found = True
                    break
                
                row += 1
            
            wb.close()
            
            if not device_found:
                self.show_no_device_found()
                
        except Exception as e:
            self.show_error_results(f"Error searching device: {str(e)}")
    
    def display_device_result(self, serial, brand, device_name, warranty, author):
        # Table header
        header_frame = ctk.CTkFrame(self.results_container, fg_color="transparent", height=45)
        header_frame.pack(fill="x", padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        headers = [("Serial", 0.2), ("Brand", 0.18), ("Device", 0.32), ("Warranty", 0.15), ("Author", 0.15)]
        left_padding = 20
        
        for i, (text, width) in enumerate(headers):
            x_pos = left_padding + sum(w * 1100 for _, w in headers[:i])
            lbl = ctk.CTkLabel(
                header_frame,
                text=text,
                font=ctk.CTkFont(size=12, weight="normal"),
                text_color="#6B7280",
                anchor="w",
                width=int(width * 1100)
            )
            lbl.place(x=x_pos, rely=0.5, anchor="w")
        
        # Separator line
        separator = ctk.CTkFrame(self.results_container, fg_color="#E5E7EB", height=1)
        separator.pack(fill="x", padx=0)
        
        # Device row
        row = ctk.CTkFrame(self.results_container, fg_color="transparent", height=50)
        row.pack(fill="x", padx=0, pady=4)
        row.pack_propagate(False)
        
        fields = [(serial, 0.2), (brand, 0.18), (device_name, 0.32), (warranty, 0.15), (author, 0.15)]
        
        for i, (value, width) in enumerate(fields):
            x_pos = left_padding + sum(w * 1100 for _, w in fields[:i])
            lbl = ctk.CTkLabel(
                row,
                text=str(value),
                font=ctk.CTkFont(size=13),
                text_color="#1F2937" if i < 2 else "#6B7280",
                anchor="w",
                width=int(width * 1100)
            )
            lbl.place(x=x_pos, rely=0.5, anchor="w")
        
        # Bottom border
        border = ctk.CTkFrame(row, fg_color="#F3F4F6", height=1)
        border.pack(side="bottom", fill="x", padx=20)
    
    def show_no_device_found(self):
        # Table header
        header_frame = ctk.CTkFrame(self.results_container, fg_color="transparent", height=45)
        header_frame.pack(fill="x", padx=0, pady=0)
        header_frame.pack_propagate(False)
        
        headers = [("Serial", 0.2), ("Brand", 0.18), ("Device", 0.32), ("Warranty", 0.15), ("Author", 0.15)]
        left_padding = 20
        
        for i, (text, width) in enumerate(headers):
            x_pos = left_padding + sum(w * 1100 for _, w in headers[:i])
            lbl = ctk.CTkLabel(
                header_frame,
                text=text,
                font=ctk.CTkFont(size=12, weight="normal"),
                text_color="#6B7280",
                anchor="w",
                width=int(width * 1100)
            )
            lbl.place(x=x_pos, rely=0.5, anchor="w")
        
        # Separator line
        separator = ctk.CTkFrame(self.results_container, fg_color="#E5E7EB", height=1)
        separator.pack(fill="x", padx=0)
        
        # No device found message
        message_frame = ctk.CTkFrame(self.results_container, fg_color="transparent")
        message_frame.pack(fill="both", expand=True, pady=40)
        
        no_device_label = ctk.CTkLabel(
            message_frame,
            text="No device found",
            font=ctk.CTkFont(size=16, weight="bold"),
            text_color="#EF4444"
        )
        no_device_label.place(relx=0.5, rely=0.5, anchor="center")
    
    def show_error_results(self, error_message):
        for widget in self.results_container.winfo_children():
            widget.destroy()
        
        error_label = ctk.CTkLabel(
            self.results_container,
            text=error_message,
            font=ctk.CTkFont(size=14),
            text_color="#EF4444"
        )
        error_label.place(relx=0.5, rely=0.5, anchor="center")
    
    # --------------- Add Mode -----------------
    def show_add_mode(self):
        self.clear_main_frame()
        self.update_active_button(self.add_btn)
        
        container = ctk.CTkFrame(self.main_frame, fg_color="white", corner_radius=12, width=500, height=400)
        container.place(relx=0.5, rely=0.5, anchor="center")
        container.pack_propagate(False)
        
        title = ctk.CTkLabel(
            container, 
            text="Add Device", 
            font=ctk.CTkFont(size=24, weight="bold"),
            text_color="#1F2937"
        )
        title.pack(pady=(40, 30))
        
        brand_frame = ctk.CTkFrame(container, fg_color="transparent")
        brand_frame.pack(pady=15, padx=40, fill="x")
        ctk.CTkLabel(
            brand_frame, 
            text="Select Brand:", 
            font=ctk.CTkFont(size=14),
            text_color="#374151"
        ).pack(anchor="w", pady=(0, 8))
        self.add_brand_selector = ctk.CTkOptionMenu(
            brand_frame, 
            values=["HP","Dell","Lenovo"], 
            width=420, 
            height=40,
            fg_color="white",
            button_color="#3B82F6",
            button_hover_color="#2563EB",
            text_color="#1F2937"
        )
        self.add_brand_selector.pack(fill="x")
        
        serial_frame = ctk.CTkFrame(container, fg_color="transparent")
        serial_frame.pack(pady=15, padx=40, fill="x")
        ctk.CTkLabel(
            serial_frame, 
            text="Serial Number:", 
            font=ctk.CTkFont(size=14),
            text_color="#374151"
        ).pack(anchor="w", pady=(0, 8))
        self.add_serial = ctk.CTkEntry(
            serial_frame, 
            placeholder_text="Enter serial number", 
            width=420, 
            height=40,
            fg_color="white",
            border_color="#E5E7EB",
            text_color="#1F2937"
        )
        self.add_serial.pack(fill="x")
        
        add_btn = ctk.CTkButton(
            container, 
            text="Add Device", 
            height=40, 
            width=420, 
            command=self.perform_add,
            fg_color="#3B82F6",
            hover_color="#2563EB",
            corner_radius=6,
            font=ctk.CTkFont(size=14, weight="bold")
        )
        add_btn.pack(pady=30, padx=40)
    
    def perform_add(self):
        brand = self.add_brand_selector.get()
        serial = self.add_serial.get()
        scrape.add_device(brand, serial, self.current_user)
        # Invalidate cache after adding
        self.cache_valid = False

if __name__ == "__main__":
    app = DeviceTicketSystem()
    app.mainloop()